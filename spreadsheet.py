from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string

import os
import sqlite3
import pandas as pd
from datetime import datetime,timedelta
import shutil
from datetime import date,time
#database connection
conn = sqlite3.connect('Face-DataBase')
c = conn.cursor()

#get current date
now=datetime.now()
current_day=date.today().weekday()
currentdate=now.strftime("%d_%m_%y")

data=pd.read_csv("time_table (copy).csv")
currenttime=now.strftime("%H:%M")

periods=data.values[current_day+1][1:]
if not os.path.exists(currentdate):
    os.makedirs(currentdate)
period_path="./"+currentdate+"/"

for i in periods:
    if not os.path.exists(period_path+i):
        os.makedirs(period_path+i)

for folder in os.listdir("./"+currentdate+"/"):
    if(os.path.exists('./'+currentdate+"/"+folder+"/"+folder+"_"+currentdate+".xlsx")):
        wb = load_workbook('./'+currentdate+"/"+folder+"/"+folder+"_"+currentdate+".xlsx")
        sheet = wb['Cse16']
        # sheet[ord() + '1']
        for col_index in range(1, 100):
            col = get_column_letter(col_index)
            if sheet.cell(row = 1, column = 1).value is None:
                col2 = get_column_letter(col_index - 1)
                # print sheet.cell('%s%s'% (col2, 1)).value
                if sheet.cell('%s%s' % (col2,1)).value != "I":
                    sheet['%s%s' % (col,1)] = "I"
                break

        #saving the file
        wb.save("./"+currentdate+"/"+folder+"/"+folder+"_"+currentdate+".xlsx")

    else:
        wb = Workbook()
        dest_filename = folder+"_"+currentdate
        #path="./"+currentdate+"/"+folder+"/"
        c.execute("SELECT * FROM Students ORDER BY Roll ASC")
        ws1 = wb.active
        ws1.title = "Cse16"
        ws1.append(('Roll Number', 'Name', 'I','II','III','IV','V','Final'))
        ws1.append(('', '', '','','','','',''))
        while True:
            a = c.fetchone()
            if a == None:
                break
            else:
                ws1.append((a[2], a[1]))
        wb.save("./"+currentdate+"/"+folder+"/"+dest_filename+".xlsx")
