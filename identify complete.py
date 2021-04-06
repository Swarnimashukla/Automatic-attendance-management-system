import pandas as pd
import numpy as np
from datetime import date,time
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import Cell
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import cv2
import shutil
import time as t
import os
import cognitive_face as CF
import global_variables as global_var
import urllib
import sqlite3
import dlib
import sys




#df1=pd.read_csv("time_table.csv")
df = pd.read_csv("time_table (copy).csv")
df




now=datetime.now()
current_day = date.today().weekday()
currentdate=now.strftime("%d_%m_%y")
dict={}
duration=0
path=""
a=np.array(df.values[0][1:])
k=0
start=[]
end=[]
periods=df.values[current_day+1][1:]
roman = ['I','II','III','IV','V']
for i in periods:
    dict[a[k]]=i
    start.append(a[k][0:5])
    end.append(a[k][6:])
    k=k+1



def is_time_between(begin_time, end_time, check_time=None):
    # If check time is not given, default to current UTC time
    check_time = check_time or datetime.now().time()
    #print(check_time)
    if begin_time < end_time:
        return check_time >= begin_time and check_time <= end_time
    else: # crosses midnight
        return check_time >= begin_time or check_time <= end_time



def path12():
    global duration
    k=0
    path=""
    for i in dict.keys():
        hs=int(start[k][0:2])
        ms=int(start[k][3:5])
        he=int(end[k][0:2])
        me=int(end[k][3:5])
        s1 = start[k]
        s2 = end[k]
        FMT = '%H:%M'
        if(is_time_between(time(hs,ms), time(he,me))):
            path="./"+currentdate+"/"+dict[i]+"/"+dict[i]+"_"+currentdate+".xlsx"
            tdelta = str(datetime.strptime(s2, FMT) - datetime.strptime(s1, FMT))
            i=int(tdelta[2:4])
            #print(i)
            duration=int(i/5)
            #print(duration)
            break
        k=k+1
    return path


def final_spreadsheet():
    df_col={}
    global currentdate
    for sub in os.listdir(currentdate):
        name = os.listdir("./"+currentdate+"/"+sub+"/")
        df_col[sub]= pd.read_excel("./"+currentdate+"/"+sub+"/"+str(name[0]),index=False)
        
    df_temp = df_col[str(sub)]

    df_temp['Roll Number']=df_col[str(sub)]['Roll Number']
    df_temp['Name']=df_col[str(sub)]['Name']
    
    for i in df_col.keys():
        df_temp[str(i)]=df_col[i]['Final']
        
    for i in os.listdir("./"+currentdate+"/"):
        shutil.rmtree("./"+currentdate+"/"+i)
    
    df_temp=df_temp.drop(["Unnamed: 0","Final"],axis=1)
    df_temp.to_excel("./"+currentdate+"/"+currentdate+"_"+"final"+".xlsx")



def final_count(path):
    global currentdate
    df= pd.read_excel(path,index=False)
    df.fillna(0,inplace=True)
    df=df.drop(0,axis=0)

    df['Final']=df['I']+df['II']+df['III']+df['IV']+df['V']
    df['Final']=np.where(df["Final"]>=3,'Present','Absent')
    df=df.drop(columns=['I','II','III','IV','V'])
    df.replace(0,np.nan,inplace=True)
    os.remove(path)
    df.to_excel(path)



def detect():

    detector = dlib.get_frontal_face_detector()

    cam = cv2.VideoCapture(0)
    detector = dlib.get_frontal_face_detector()
    ret, frame = cam.read()
    ret, frame = cam.read()
    ret, frame = cam.read()
    count=1
    if not os.path.exists('./pics'):
        os.makedirs("./pics")
    cv2.imwrite("./pics/framee%d.jpg" % count , frame)
    cam.release()

    img = cv2.imread('./pics/framee1.jpg')
    dets = detector(img, 1)
    if not os.path.exists('./Cropped_faces'):
        os.makedirs('./Cropped_faces')
        print("detected = {}".format(len(dets)))
        for i, d in enumerate(dets):
            cv2.imwrite('./Cropped_faces/face' + str(i + 1) + '.jpg', img[d.top():d.bottom(), d.left():d.right()])


def identify(path,column):
    #get current date
    global currentdate
    wb = load_workbook(path)
    sheet = wb['Cse16']

    def getColumn():
        for i in range(1, len(list(sheet.rows)[0]) + 1):
            col = get_column_letter(i)
            if sheet['%s%s'% (col,'1')].value == column:
                return col

    Key = global_var.key

    CF.Key.set(Key)

    BASE_URL = global_var.BASE_URL  # Replace with your regional Base URL
    CF.BaseUrl.set(BASE_URL)

    connect = sqlite3.connect("Face-DataBase")
    #c = connect.cursor()

    attend = [0 for i in range(60)]

    currentDir = os.path.dirname(os.path.abspath(''))+"/"+"test"
    directory = os.path.join(currentDir, 'Cropped_faces')
    
    
    for filename in os.listdir(directory):
        if filename.endswith(".jpg"):
            imgurl = urllib.request.pathname2url(os.path.join(directory, filename))
            #imgurl = imgurl[3:]
            print("imgurl = {}".format(imgurl))
            res = CF.face.detect(imgurl)
            print("Res = {}".format(res))

            if len(res) < 1:
                print("No face detected.")
                continue

            faceIds = []
            for face in res:
                faceIds.append(face['faceId'])
            res = CF.face.identify(faceIds, global_var.personGroupId)
            print(filename)
            print("res = {}".format(res))

            for face  in res:
                if not face['candidates']:
                    print("Unknown")
                else:
                    personId = face['candidates'][0]['personId']
                    print("personid = {}".format(personId))
                    #cmd =  + personId
                    cur = connect.execute("SELECT * FROM Students WHERE personID = (?)", (personId,))
                    #print("cur = {}".format(cur))
                    for row in cur:
                        print("aya")
                        print("row = {}".format(row))
                        attend[int(row[0])] += 1
                    print("---------- " + row[1] + " recognized ----------")
            t.sleep(6)

    for row in range(2, len(list(sheet.columns)[0]) + 1):
        rn = sheet.cell(row = row, column  =1).value
        if rn is not None:
            print("rn = {}".format(rn))
            rn = rn[-2:]
            if attend[int(rn)] != 0:
                col = getColumn()
                print("col = {}".format(col))
                sheet['%s%s' % (col, str(row))] = 1

    wb.save(path)


os.system("python3 spreadsheet.py")


while(1):
    if((current_day!=date.today().weekday()) or (currentdate!=now.strftime("%d_%m_%y"))):
        current_day=date.today().weekday()
        currentdate=now.strftime("%d_%m_%y")
        dict={}
        a=np.array(df.values[0][1:])
        k=0
        periods=df.values[current_day+1][1:]
        for i in periods:
            dict[a[k]]=i
            start.append(a[k][0:5])
            end.append(a[k][6:])
            k=k+1
        os.system("python3 spreadsheet.py")
        
    while is_time_between(time(int(start[0][0:2]),int(start[0][3:5])), time(int(end[len(end)-1][0:2]),int(end[len(end)-1][3:5]))):
        hello12=datetime.now().time().strftime("%H:%M")
        curr = datetime.now()
        #print(duration)
        path=path12()
        if hello12 in start:
            interval = []    
            print(hello12)
            for x in range(5):
                curr = curr + timedelta(minutes = duration)
                interval.append(curr.strftime("%H:%M"))
            print(interval)
            t.sleep(60)
            
            
        if hello12 in interval:
            #print(hello12)
            path=path12()
            index = interval.index(hello12)
            column12 = roman[index]
            #call detect to take photos
            detect()
            #call identify with the path and the name of the column
            identify(path,column12)
            if column12=="V":
                final_count(path)
            t.sleep(60)
            
        if hello12==end[len(end)-1]:
            final_spreadsheet()


currentDir = os.path.dirname(os.path.abspath(''))+"/"+"final_project"
directory = os.path.join(currentDir, 'Cropped_faces')
print(currentDir)
print(type(currentDir))
print(directory)
print(type(directory))
print(os.path.dirname(os.path.abspath('')))

